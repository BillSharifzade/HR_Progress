#!/usr/bin/env python3
"""Convert the digital-profile questionnaire export to JSON.

The Go importer does the matching, parsing and writing; this script only turns
the xlsx into a stable, reviewable intermediate so the backend needs no Excel
dependency (adding one dragged golang.org/x/crypto forward by 25 minor
versions, which is not a trade worth making to read a spreadsheet).

Column order is positional and asserted against the header text, so a
re-exported file with shuffled or renamed columns fails loudly instead of
importing the wrong answers into the wrong fields.

Usage:  form_to_json.py <input.xlsx> <output.json>
"""
import json
import re
import sys

import openpyxl

# (index, key, expected header fragment). The fragment is matched loosely —
# the export carries stray leading/trailing spaces and newlines in headers.
COLUMNS = [
    (0,  "submitted_at",                "Отметка времени"),
    (1,  "full_name",                   "ФИО"),
    (2,  "department",                  "Департамент"),
    (3,  "education_levels",            "основное образование"),
    (4,  "institution",                 "учебное заведение"),
    (5,  "specialty",                   "специальность"),
    (6,  "certificates",                "сертификаты"),
    (7,  "lang_tajik",                  "[Таджикский]"),
    (8,  "lang_russian",                "[Русский]"),
    (9,  "lang_english",                "[Английский]"),
    (10, "lang_chinese",                "[Китайский]"),
    (11, "lang_german",                 "[Немецкий]"),
    (12, "lang_turkish",                "[Турецкий]"),
    (13, "prior_experience_band",       "профессиональный стаж"),
    (14, "notable_projects",            "значимых проектах"),
    (15, "previous_employers",          "компаниях Вы работали"),
    (16, "career_goal",                 "карьерная цель"),
    (17, "development_directions",      "направлениях Вы хотели бы развиваться"),
    (18, "mobility_readiness",          "переход в другой Департамент"),
    (19, "relocation_readiness",        "релокации"),
    (20, "internal_projects_readiness", "внутренних проектах"),
    (21, "expertise_areas",             "наиболее компетентным"),
    (22, "colleagues_ask_about",        "обращаются коллеги"),
    (23, "teaching_readiness",          "внутреннее обучение"),
    (24, "teaching_topics",             "каким темам"),
    (25, "hobbies",                     "хобби"),
    (26, "professional_interests",      "профессиональные темы"),
    (27, "learning_formats",            "формат обучения"),
    (28, "learning_hours_band",         "Сколько времени"),
    (29, "extra_note",                  "дополнительная информация"),
]


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def main():
    if len(sys.argv) != 3:
        print(__doc__, file=sys.stderr)
        return 2
    src, dst = sys.argv[1], sys.argv[2]

    ws = openpyxl.load_workbook(src, data_only=True).active
    header = [norm(c.value) for c in ws[1]]

    for idx, key, fragment in COLUMNS:
        if idx >= len(header):
            raise SystemExit(f"column {idx} ({key}) missing: sheet has {len(header)} columns")
        if fragment.lower() not in header[idx].lower():
            raise SystemExit(
                f"column {idx} should be {key!r} (expected to contain {fragment!r}) "
                f"but the header reads {header[idx]!r} — refusing to guess"
            )

    out = []
    for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        rec = {"source_row": n}
        for idx, key, _ in COLUMNS:
            value = row[idx] if idx < len(row) else None
            if key == "submitted_at":
                rec[key] = value.isoformat() if hasattr(value, "isoformat") else norm(value)
            else:
                rec[key] = norm(value)
        out.append(rec)

    with open(dst, "w", encoding="utf-8") as f:
        json.dump({"question_texts": {k: header[i] for i, k, _ in COLUMNS}, "rows": out},
                  f, ensure_ascii=False, indent=1)
    print(f"wrote {len(out)} rows to {dst}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
